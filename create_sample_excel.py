import openpyxl
from openpyxl.styles import Font, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '選手一覧'

# ヘッダー
headers = ['背番号', '氏名', 'フリガナ', '学年', 'ポジション', '身長', '前所属']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center')

# サンプルデータ
players = [
    (1, '山田太郎', 'ヤマダタロウ', 3, 'GK', 178, 'FCジュニア'),
    (2, '鈴木次郎', 'スズキジロウ', 3, 'DF', 175, 'サッカークラブA'),
    (3, '佐藤三郎', 'サトウサブロウ', 3, 'DF', 172, 'FCユース'),
    (4, '田中四郎', 'タナカシロウ', 2, 'DF', 170, 'ジュニアユース'),
    (5, '高橋五郎', 'タカハシゴロウ', 2, 'DF', 168, 'SCジュニア'),
    (6, '伊藤六郎', 'イトウロクロウ', 2, 'MF', 165, 'FC東'),
    (7, '渡辺七郎', 'ワタナベシチロウ', 3, 'MF', 173, 'サッカー少年団'),
    (8, '中村八郎', 'ナカムラハチロウ', 2, 'MF', 168, 'FCセントラル'),
    (9, '小林九郎', 'コバヤシクロウ', 3, 'FW', 175, 'SCユース'),
    (10, '加藤十郎', 'カトウジュウロウ', 3, 'MF', 170, 'アカデミー'),
    (11, '吉田一', 'ヨシダハジメ', 2, 'FW', 172, 'FCウエスト'),
    (12, '山本二郎', 'ヤマモトジロウ', 1, 'GK', 176, 'ジュニアFC'),
    (13, '松本三郎', 'マツモトサブロウ', 1, 'DF', 167, 'SC南'),
    (14, '井上四郎', 'イノウエシロウ', 1, 'MF', 164, 'FCノース'),
    (15, '木村五郎', 'キムラゴロウ', 1, 'FW', 169, 'サッカークラブB'),
    (16, '石田六郎', 'イシダロクロウ', 2, 'DF', 171, 'ユースFC'),
    (17, '斎藤七郎', 'サイトウシチロウ', 2, 'MF', 166, 'アカデミーJr'),
    (18, '清水八郎', 'シミズハチロウ', 1, 'MF', 163, 'FCイースト'),
    (19, '森田九郎', 'モリタクロウ', 1, 'FW', 170, 'SCセントラル'),
    (20, '池田十郎', 'イケダジュウロウ', 1, 'DF', 165, 'ジュニアSC'),
]

for row, player in enumerate(players, 2):
    for col, value in enumerate(player, 1):
        ws.cell(row=row, column=col, value=value)

# 列幅調整
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 8
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 8
ws.column_dimensions['G'].width = 18

wb.save('D:/UrawaCup2/sample_players.xlsx')
print('Excel file created: D:/UrawaCup2/sample_players.xlsx')
