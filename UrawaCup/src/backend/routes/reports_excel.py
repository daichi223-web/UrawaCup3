"""
Reports Excel Export - 報告書Excel出力ルーター
特別レポートのExcel出力エンドポイントを提供
"""

import io
from datetime import date
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from database import get_db
from models.tournament import Tournament
from models.match import Match, MatchStage, MatchStatus
from models.group import Group
from models.standing import Standing
from models.team import Team

router = APIRouter()


@router.get("/export/group-standings/excel")
def export_group_standings_excel(
    tournament_id: int = Query(..., description="大会ID"),
    group_id: Optional[str] = Query(None, description="グループID"),
    db: Session = Depends(get_db),
):
    """グループ順位表をExcelで出力"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="大会が見つかりません")

    query = db.query(Standing).filter(Standing.tournament_id == tournament_id)
    if group_id:
        query = query.filter(Standing.group_id == group_id)
    standings = query.order_by(Standing.group_id, Standing.rank).all()

    if not standings:
        raise HTTPException(status_code=404, detail="順位データが見つかりません")

    wb = Workbook()
    ws = wb.active
    ws.title = "グループ順位表"

    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    ws.merge_cells('A1:J1')
    ws['A1'] = f"{tournament.name} グループ順位表"
    ws['A1'].font = header_font

    row = 3
    current_group = None

    for standing in standings:
        if standing.group_id != current_group:
            if current_group is not None:
                row += 1

            current_group = standing.group_id
            group = db.query(Group).filter(Group.id == standing.group_id).first()
            group_name = group.name if group else standing.group_id

            ws[f'A{row}'] = f"【{group_name}】"
            ws[f'A{row}'].font = subheader_font
            row += 1

            headers = ['順位', 'チーム', '試合', '勝', '分', '負', '得点', '失点', '得失点差', '勝点']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
            row += 1

        team = db.query(Team).filter(Team.id == standing.team_id).first()
        team_name = team.short_name or team.name if team else "Unknown"

        values = [
            standing.rank,
            team_name,
            standing.played,
            standing.won,
            standing.drawn,
            standing.lost,
            standing.goals_for,
            standing.goals_against,
            standing.goal_difference,
            standing.points,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 6

    row += 2
    ws.cell(row=row, column=1, value="浦和カップ運営事務局")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"group_standings_{tournament_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/final-day-schedule/excel")
def export_final_day_schedule_excel(
    tournament_id: int = Query(..., description="大会ID"),
    target_date: date = Query(..., description="対象日"),
    db: Session = Depends(get_db),
):
    """最終日組み合わせ表をExcelで出力"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    matches = (
        db.query(Match)
        .options(joinedload(Match.home_team), joinedload(Match.away_team), joinedload(Match.venue))
        .filter(Match.tournament_id == tournament_id, Match.match_date == target_date)
        .order_by(Match.venue_id, Match.match_order)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "最終日組み合わせ"

    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    date_str = target_date.strftime("%m月%d日")
    weekday = ["月", "火", "水", "木", "金", "土", "日"][target_date.weekday()]

    ws.merge_cells('A1:E1')
    ws['A1'] = f"最終日組み合わせ表 - {date_str}（{weekday}）"
    ws['A1'].font = header_font

    venue_matches = {}
    knockout_matches = []
    for match in matches:
        if match.stage in [MatchStage.SEMIFINAL, MatchStage.THIRD_PLACE, MatchStage.FINAL]:
            knockout_matches.append(match)
        else:
            venue_id = match.venue_id or 0
            if venue_id not in venue_matches:
                venue_matches[venue_id] = {'venue': match.venue, 'matches': []}
            venue_matches[venue_id]['matches'].append(match)

    row = 3
    ws[f'A{row}'] = "【順位リーグ】"
    ws[f'A{row}'].font = subheader_font
    row += 2

    for venue_id, data in venue_matches.items():
        venue = data['venue']
        ws[f'A{row}'] = f"● {venue.name if venue else '不明'}"
        ws[f'A{row}'].font = subheader_font
        row += 1

        for col, header in enumerate(['No', 'KO', 'ホーム', 'vs', 'アウェイ'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font_white
            cell.border = thin_border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1

        for match in data['matches']:
            home = match.home_team.short_name or match.home_team.name if match.home_team else "TBD"
            away = match.away_team.short_name or match.away_team.name if match.away_team else "TBD"
            ko_time = match.match_time.strftime("%H:%M") if match.match_time else ""
            for col, value in enumerate([match.match_order, ko_time, home, 'vs', away], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1
        row += 1

    if knockout_matches:
        row += 1
        ws[f'A{row}'] = "【3決・決勝戦】"
        ws[f'A{row}'].font = subheader_font
        row += 2

        venue = knockout_matches[0].venue if knockout_matches else None
        ws[f'A{row}'] = f"● {venue.name if venue else '駒場スタジアム'}"
        ws[f'A{row}'].font = subheader_font
        row += 1

        knockout_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for col, header in enumerate(['種別', 'KO', 'ホーム', 'vs', 'アウェイ'], 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True, size=10)
            cell.border = thin_border
            cell.fill = knockout_fill
            cell.alignment = Alignment(horizontal='center')
        row += 1

        type_map = {MatchStage.SEMIFINAL: "準決勝", MatchStage.THIRD_PLACE: "3位決定戦", MatchStage.FINAL: "決勝"}
        for match in knockout_matches:
            home = match.home_team.short_name or match.home_team.name if match.home_team else "TBD"
            away = match.away_team.short_name or match.away_team.name if match.away_team else "TBD"
            ko_time = match.match_time.strftime("%H:%M") if match.match_time else ""
            for col, value in enumerate([type_map.get(match.stage, ""), ko_time, home, 'vs', away], 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
            row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 18

    row += 2
    ws.cell(row=row, column=1, value="浦和カップ運営事務局")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"final_day_schedule_{target_date.isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/final-result/excel")
def export_final_result_excel(
    tournament_id: int = Query(..., description="大会ID"),
    db: Session = Depends(get_db),
):
    """最終結果報告書をExcelで出力"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="大会が見つかりません")

    knockout_matches = (
        db.query(Match)
        .options(joinedload(Match.home_team), joinedload(Match.away_team))
        .filter(
            Match.tournament_id == tournament_id,
            Match.stage.in_([MatchStage.SEMIFINAL, MatchStage.THIRD_PLACE, MatchStage.FINAL])
        )
        .order_by(Match.stage, Match.match_order)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "最終結果"

    header_font = Font(bold=True, size=16)
    subheader_font = Font(bold=True, size=12)
    normal_font = Font(size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    gold_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    silver_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    bronze_fill = PatternFill(start_color="CD7F32", end_color="CD7F32", fill_type="solid")

    ws.merge_cells('A1:E1')
    ws['A1'] = f"{tournament.name} 最終結果"
    ws['A1'].font = header_font

    row = 3

    ws[f'A{row}'] = "【決勝トーナメント結果】"
    ws[f'A{row}'].font = subheader_font
    row += 2

    type_map = {MatchStage.SEMIFINAL: "準決勝", MatchStage.THIRD_PLACE: "3位決定戦", MatchStage.FINAL: "決勝"}

    for col, header in enumerate(['種別', 'ホーム', 'スコア', 'アウェイ', '結果'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=10)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    for match in knockout_matches:
        home = match.home_team.short_name or match.home_team.name if match.home_team else "TBD"
        away = match.away_team.short_name or match.away_team.name if match.away_team else "TBD"

        if match.status == MatchStatus.COMPLETED:
            h_total = match.home_score_total or 0
            a_total = match.away_score_total or 0
            score = f"{h_total} - {a_total}"
            if h_total > a_total:
                result = f"{home} 勝利"
            elif h_total < a_total:
                result = f"{away} 勝利"
            else:
                if match.home_pk is not None and match.away_pk is not None:
                    score += f" (PK {match.home_pk}-{match.away_pk})"
                    result = f"{home if match.home_pk > match.away_pk else away} 勝利"
                else:
                    result = "引分"
        else:
            score = "- vs -"
            result = "未実施"

        values = [type_map.get(match.stage, ""), home, score, away, result]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    row += 2
    ws[f'A{row}'] = "【最終順位】"
    ws[f'A{row}'].font = subheader_font
    row += 2

    final_match = next((m for m in knockout_matches if m.stage == MatchStage.FINAL and m.status == MatchStatus.COMPLETED), None)
    third_place_match = next((m for m in knockout_matches if m.stage == MatchStage.THIRD_PLACE and m.status == MatchStatus.COMPLETED), None)

    rankings = []
    if final_match:
        h_total = final_match.home_score_total or 0
        a_total = final_match.away_score_total or 0
        h_pk = final_match.home_pk or 0
        a_pk = final_match.away_pk or 0
        if h_total > a_total or (h_total == a_total and h_pk > a_pk):
            rankings.append(("優勝", final_match.home_team))
            rankings.append(("準優勝", final_match.away_team))
        else:
            rankings.append(("優勝", final_match.away_team))
            rankings.append(("準優勝", final_match.home_team))

    if third_place_match:
        h_total = third_place_match.home_score_total or 0
        a_total = third_place_match.away_score_total or 0
        h_pk = third_place_match.home_pk or 0
        a_pk = third_place_match.away_pk or 0
        if h_total > a_total or (h_total == a_total and h_pk > a_pk):
            rankings.append(("3位", third_place_match.home_team))
        else:
            rankings.append(("3位", third_place_match.away_team))

    fills = [gold_fill, silver_fill, bronze_fill]
    for i, (rank, team) in enumerate(rankings):
        team_name = team.short_name or team.name if team else "TBD"
        cell_rank = ws.cell(row=row, column=1, value=rank)
        cell_team = ws.cell(row=row, column=2, value=team_name)
        cell_rank.font = Font(bold=True, size=12)
        cell_team.font = Font(bold=True, size=12)
        cell_rank.border = thin_border
        cell_team.border = thin_border
        if i < len(fills):
            cell_rank.fill = fills[i]
            cell_team.fill = fills[i]
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15

    row += 2
    ws.cell(row=row, column=1, value="浦和カップ運営事務局")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"final_result_{tournament_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
