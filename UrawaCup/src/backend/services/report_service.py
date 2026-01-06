"""
ReportService - 報告書生成サービス

PDF/Excel形式で試合報告書を生成する

サブモジュール:
- services.reports: 報告書ジェネレータ（TournaMate_Report_Formats.md準拠）
"""

import io
from datetime import date
from typing import Optional, List
from sqlalchemy.orm import Session, joinedload

from models.tournament import Tournament
from models.match import Match, MatchStage, MatchStatus
from models.venue import Venue
from models.goal import Goal

# 新しいサブモジュールからインポート
from services.reports import DailyReportGenerator, SenderInfo


class ReportService:
    """報告書生成サービス"""

    def __init__(self, db: Session):
        self.db = db

    def _get_matches(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> List[Match]:
        """報告書用の試合データを取得"""
        query = (
            self.db.query(Match)
            .options(
                joinedload(Match.home_team),
                joinedload(Match.away_team),
                joinedload(Match.venue),
                joinedload(Match.goals).joinedload(Goal.team),
            )
            .filter(
                Match.tournament_id == tournament_id,
                Match.match_date == target_date,
                Match.status == MatchStatus.COMPLETED,
                Match.stage != MatchStage.TRAINING,  # 研修試合は除外
            )
        )

        if venue_id:
            query = query.filter(Match.venue_id == venue_id)

        return query.order_by(Match.venue_id, Match.match_order).all()

    def _get_tournament(self, tournament_id: int) -> Tournament:
        """大会情報を取得"""
        tournament = self.db.query(Tournament).filter(Tournament.id == tournament_id).first()
        if not tournament:
            raise ValueError(f"大会が見つかりません (ID: {tournament_id})")
        return tournament

    def generate_pdf(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        PDF形式の報告書を生成

        ReportLabを使用してPDFを生成
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            # ReportLabがインストールされていない場合はシンプルなPDFを生成
            return self._generate_simple_pdf(tournament_id, target_date, venue_id)

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        buffer = io.BytesIO()

        # PDF作成
        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # 日本語フォントの登録（フォールバック用）
        try:
            # Windowsの場合
            pdfmetrics.registerFont(TTFont('Gothic', 'C:/Windows/Fonts/msgothic.ttc'))
            font_name = 'Gothic'
        except:
            try:
                # Macの場合
                pdfmetrics.registerFont(TTFont('Gothic', '/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc'))
                font_name = 'Gothic'
            except:
                font_name = 'Helvetica'

        # ヘッダー
        c.setFont(font_name, 16)
        c.drawString(30 * mm, height - 20 * mm, f"{tournament.name}")
        c.setFont(font_name, 12)
        c.drawString(30 * mm, height - 30 * mm, f"試合結果報告書 - {target_date.isoformat()}")

        if venue_id:
            venue = self.db.query(Venue).filter(Venue.id == venue_id).first()
            if venue:
                c.drawString(30 * mm, height - 40 * mm, f"会場: {venue.name}")

        # 試合情報
        y_position = height - 60 * mm
        c.setFont(font_name, 10)

        for match in matches:
            if y_position < 40 * mm:
                c.showPage()
                y_position = height - 20 * mm
                c.setFont(font_name, 10)

            # 試合情報
            home = match.home_team.short_name or match.home_team.name
            away = match.away_team.short_name or match.away_team.name
            time_str = match.match_time.strftime("%H:%M")

            h1 = match.home_score_half1 or 0
            h2 = match.home_score_half2 or 0
            a1 = match.away_score_half1 or 0
            a2 = match.away_score_half2 or 0

            c.drawString(30 * mm, y_position, f"第{match.match_order}試合 {time_str}")
            y_position -= 6 * mm

            score_line = f"{home} {h1+h2} - {a1+a2} {away}"
            c.drawString(35 * mm, y_position, score_line)
            y_position -= 5 * mm

            c.drawString(35 * mm, y_position, f"(前半 {h1}-{a1}、後半 {h2}-{a2})")

            if match.has_penalty_shootout:
                y_position -= 5 * mm
                c.drawString(35 * mm, y_position, f"PK {match.home_pk}-{match.away_pk}")

            y_position -= 5 * mm

            # 得点経過
            for goal in sorted(match.goals, key=lambda g: (g.half, g.minute)):
                half_text = "前半" if goal.half == 1 else "後半"
                goal_text = f"{half_text}{goal.minute}分 {goal.team.name} {goal.player_name}"
                if goal.is_own_goal:
                    goal_text += "(OG)"
                if goal.is_penalty:
                    goal_text += "(PK)"

                c.drawString(40 * mm, y_position, goal_text)
                y_position -= 4 * mm

            y_position -= 8 * mm

        # フッター
        c.setFont(font_name, 8)
        c.drawString(30 * mm, 15 * mm, "浦和カップ運営事務局")

        c.save()
        buffer.seek(0)

        return buffer

    def _generate_simple_pdf(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        シンプルなテキストベースのPDFを生成（ReportLabがない場合）
        """
        # 簡易的なPDF生成（実際の運用ではReportLabをインストール推奨）
        buffer = io.BytesIO()

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        # PDFヘッダー
        pdf_content = b"%PDF-1.4\n"
        pdf_content += b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        pdf_content += b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        pdf_content += b"3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R/Resources<<>>/Contents 4 0 R>>endobj\n"

        # コンテンツ
        text = f"Match Report - {target_date.isoformat()}\n\n"
        for match in matches:
            home = match.home_team.name
            away = match.away_team.name
            score = f"{match.home_score_total or 0}-{match.away_score_total or 0}"
            text += f"{match.match_order}. {home} {score} {away}\n"

        stream = f"4 0 obj<</Length {len(text)}>>stream\n{text}\nendstream endobj\n"
        pdf_content += stream.encode('latin-1', errors='replace')
        pdf_content += b"xref\n0 5\n0000000000 65535 f \n"
        pdf_content += b"trailer<</Size 5/Root 1 0 R>>\nstartxref\n0\n%%EOF"

        buffer.write(pdf_content)
        buffer.seek(0)

        return buffer

    def generate_excel(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        Excel形式の報告書を生成

        openpyxlを使用してExcelファイルを生成
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxlがインストールされていません。pip install openpyxl を実行してください。")

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "試合結果"

        # スタイル設定
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー
        ws.merge_cells('A1:G1')
        ws['A1'] = tournament.name
        ws['A1'].font = header_font

        ws.merge_cells('A2:G2')
        ws['A2'] = f"試合結果報告書 - {target_date.isoformat()}"
        ws['A2'].font = subheader_font

        if venue_id:
            venue = self.db.query(Venue).filter(Venue.id == venue_id).first()
            if venue:
                ws.merge_cells('A3:G3')
                ws['A3'] = f"会場: {venue.name}"

        # テーブルヘッダー
        row = 5
        headers = ['試合', '時刻', 'ホーム', '前半', '後半', '合計', 'アウェイ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # 列幅設定
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 20

        # 試合データ
        for match in matches:
            row += 1

            home = match.home_team.short_name or match.home_team.name
            away = match.away_team.short_name or match.away_team.name
            time_str = match.match_time.strftime("%H:%M")

            h1 = match.home_score_half1 or 0
            h2 = match.home_score_half2 or 0
            a1 = match.away_score_half1 or 0
            a2 = match.away_score_half2 or 0

            values = [
                match.match_order,
                time_str,
                home,
                f"{h1}-{a1}",
                f"{h2}-{a2}",
                f"{h1+h2}-{a1+a2}",
                away,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            # PK戦
            if match.has_penalty_shootout:
                row += 1
                ws.cell(row=row, column=6, value=f"PK {match.home_pk}-{match.away_pk}")

            # 得点経過
            for goal in sorted(match.goals, key=lambda g: (g.half, g.minute)):
                row += 1
                half_text = "前半" if goal.half == 1 else "後半"
                goal_text = f"{half_text}{goal.minute}分 {goal.team.name} {goal.player_name}"
                if goal.is_own_goal:
                    goal_text += "(OG)"
                if goal.is_penalty:
                    goal_text += "(PK)"

                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
                ws.cell(row=row, column=3, value=goal_text).font = normal_font

            row += 1  # 空行

        # フッター
        row += 2
        ws.cell(row=row, column=1, value="浦和カップ運営事務局")

        # ファイルに出力
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def generate_daily_report_v2(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None,
        sender_org: Optional[str] = None,
        sender_name: Optional[str] = None,
        sender_contact: Optional[str] = None,
    ) -> io.BytesIO:
        """
        日次試合結果報告書を生成（新フォーマット）

        TournaMate_Report_Formats.md に準拠した新しいレイアウトで生成

        Args:
            tournament_id: 大会ID
            target_date: 対象日
            venue_id: 会場ID（省略時は全会場）
            sender_org: 発信元所属
            sender_name: 発信元氏名
            sender_contact: 発信元連絡先

        Returns:
            生成されたPDFのBytesIO
        """
        sender = None
        if sender_org and sender_name:
            sender = SenderInfo(
                organization=sender_org,
                name=sender_name,
                contact=sender_contact or "",
            )

        generator = DailyReportGenerator(
            db=self.db,
            tournament_id=tournament_id,
            target_date=target_date,
            venue_id=venue_id,
            sender=sender,
        )
        return generator.generate()
