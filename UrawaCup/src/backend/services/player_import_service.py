"""
選手インポートサービス

浦和カップ参加申込書（Excel）のパースとインポート処理
仕様書: D:/UrawaCup/Requirement/PlayerManagement_Module_Spec.md
"""

import re
import io
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, field
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models.player import Player, normalize_name
from models.staff import Staff
from models.team_uniform import TeamUniform


@dataclass
class ParsedTeamInfo:
    """チーム基本情報"""
    name: Optional[str] = None
    postal_code: Optional[str] = None
    address: Optional[str] = None
    tel: Optional[str] = None
    fax: Optional[str] = None


@dataclass
class ParsedStaff:
    """パース済みスタッフ情報"""
    role: str
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None


@dataclass
class ParsedPlayer:
    """パース済み選手情報"""
    number: Optional[int] = None
    name: str = ""
    name_kana: Optional[str] = None
    position: Optional[str] = None
    height: Optional[int] = None
    grade: Optional[int] = None
    previous_team: Optional[str] = None
    row_number: int = 0


@dataclass
class ParsedUniform:
    """パース済みユニフォーム情報"""
    player_type: str  # GK / FP
    uniform_type: str  # primary / secondary
    shirt_color: Optional[str] = None
    pants_color: Optional[str] = None
    socks_color: Optional[str] = None


@dataclass
class ParseResult:
    """パース結果"""
    format: str = "unknown"
    team_info: Optional[ParsedTeamInfo] = None
    staff: List[ParsedStaff] = field(default_factory=list)
    uniforms: List[ParsedUniform] = field(default_factory=list)
    players: List[ParsedPlayer] = field(default_factory=list)
    errors: List[Dict[str, Any]] = field(default_factory=list)
    warnings: List[Dict[str, Any]] = field(default_factory=list)


class UrawaCupExcelParser:
    """浦和カップ参加申込書パーサー"""

    # セクション検出用キーワード
    SECTION_KEYWORDS = {
        'team_info': ['チーム', '学校', '名'],
        'staff': ['監督', 'コーチ', '携帯', 'メール', '帯同審判'],
        'uniform': ['ユニフォーム', 'ユニホーム', 'シャツ', 'GK', 'FP'],
        'players': ['登録メンバー', 'ポジション', 'Pos', '背番号', '氏名'],
    }

    # ポジション正規化マッピング
    POSITION_MAP = {
        'GK': 'GK', 'ゴールキーパー': 'GK', 'G': 'GK',
        'DF': 'DF', 'ディフェンダー': 'DF', 'D': 'DF',
        'MF': 'MF', 'ミッドフィルダー': 'MF', 'M': 'MF',
        'FW': 'FW', 'フォワード': 'FW', 'F': 'FW',
    }

    def parse(self, file_content: bytes, filename: str = "") -> ParseResult:
        """Excelファイルを解析"""
        result = ParseResult()

        try:
            # Excelファイルを読み込み
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active

            # セクション位置を検出
            sections = self._detect_sections(ws)

            # 各セクションをパース
            if 'team_info' in sections:
                result.team_info = self._parse_team_info(ws, sections['team_info'])

            if 'staff' in sections:
                result.staff = self._parse_staff(ws, sections['staff'])

            if 'uniform' in sections:
                result.uniforms = self._parse_uniforms(ws, sections['uniform'])

            if 'players' in sections:
                result.players = self._parse_players(ws, sections['players'])
                result.format = "urawa_cup_application"
            else:
                # プレイヤーセクションが見つからない場合、シンプル形式を試行
                result.players = self._parse_simple_format(ws)
                result.format = "simple" if result.players else "unknown"

            wb.close()

        except Exception as e:
            result.errors.append({
                "row": 0,
                "field": "file",
                "type": "error",
                "message": f"Excelファイルの読み込みに失敗しました: {str(e)}"
            })

        return result

    def _detect_sections(self, ws: Worksheet) -> Dict[str, int]:
        """キーワードでセクション開始位置を検出"""
        sections = {}

        for row_idx in range(1, min(ws.max_row + 1, 100)):  # 最大100行まで探索
            row_text = self._get_row_text(ws, row_idx)

            for section_name, keywords in self.SECTION_KEYWORDS.items():
                if section_name not in sections:
                    if any(kw in row_text for kw in keywords):
                        sections[section_name] = row_idx

        return sections

    def _get_row_text(self, ws: Worksheet, row_idx: int) -> str:
        """行のテキストを結合して取得"""
        cells = []
        for col in range(1, min(ws.max_column + 1, 20)):
            cell = ws.cell(row=row_idx, column=col)
            if cell.value:
                cells.append(str(cell.value))
        return ' '.join(cells)

    def _parse_team_info(self, ws: Worksheet, start_row: int) -> ParsedTeamInfo:
        """チーム基本情報を抽出"""
        info = ParsedTeamInfo()

        # 開始行から数行を探索
        for row_idx in range(start_row, min(start_row + 10, ws.max_row + 1)):
            row_text = self._get_row_text(ws, row_idx)

            # チーム名
            if 'チーム' in row_text or '学校' in row_text:
                # 隣のセルからチーム名を取得
                for col in range(1, 10):
                    cell = ws.cell(row=row_idx, column=col)
                    if cell.value and 'チーム' not in str(cell.value) and '学校' not in str(cell.value):
                        potential_name = str(cell.value).strip()
                        if len(potential_name) > 2:
                            info.name = potential_name
                            break

            # 電話番号
            tel_match = re.search(r'TEL[：:\s]*([0-9\-]+)', row_text)
            if tel_match:
                info.tel = tel_match.group(1)

            # FAX
            fax_match = re.search(r'FAX[：:\s]*([0-9\-]+)', row_text)
            if fax_match:
                info.fax = fax_match.group(1)

            # 郵便番号
            postal_match = re.search(r'[〒〶]?([0-9]{3}[-ー]?[0-9]{4})', row_text)
            if postal_match:
                info.postal_code = postal_match.group(1)

        return info

    def _parse_staff(self, ws: Worksheet, start_row: int) -> List[ParsedStaff]:
        """スタッフ情報を抽出"""
        staff_list = []
        current_phone = None
        current_email = None

        for row_idx in range(start_row, min(start_row + 15, ws.max_row + 1)):
            row_text = self._get_row_text(ws, row_idx)

            # 携帯電話
            phone_match = re.search(r'携帯[電話]*[：:\s]*([0-9\-]+)', row_text)
            if phone_match:
                current_phone = phone_match.group(1)

            # メールアドレス
            email_match = re.search(r'[\w\.\-]+@[\w\.\-]+\.\w+', row_text)
            if email_match:
                current_email = email_match.group(0)

            # 監督名
            if '監督' in row_text and '名' in row_text:
                kantoku_match = re.search(r'監督名?[：:\s]*([^\s携帯コーチ]+)', row_text)
                if kantoku_match:
                    name = kantoku_match.group(1).strip()
                    if name and len(name) >= 2:
                        staff_list.append(ParsedStaff(
                            role="監督",
                            name=name,
                            phone=current_phone,
                            email=current_email
                        ))

            # コーチ名
            if 'コーチ' in row_text:
                coach_match = re.search(r'コーチ名?[：:\s]*(.+?)(?=メール|帯同|$)', row_text)
                if coach_match:
                    coaches_text = coach_match.group(1)
                    # 全角・半角スペースで分割
                    coach_names = re.split(r'[\s　]+', coaches_text)
                    for name in coach_names:
                        name = name.strip()
                        if name and len(name) >= 2:
                            staff_list.append(ParsedStaff(role="コーチ", name=name))

            # 帯同審判
            if '帯同審判' in row_text:
                referee_match = re.search(r'帯同審判[員]?[：:\s]*([^\s]+)', row_text)
                if referee_match:
                    name = referee_match.group(1).strip()
                    if name and len(name) >= 2:
                        staff_list.append(ParsedStaff(role="帯同審判", name=name))

        return staff_list

    def _parse_uniforms(self, ws: Worksheet, start_row: int) -> List[ParsedUniform]:
        """ユニフォーム情報を抽出"""
        uniforms = []

        # ヘッダー行を探す（正/副の列位置を特定）
        header_row = None
        for row_idx in range(start_row, min(start_row + 5, ws.max_row + 1)):
            row_text = self._get_row_text(ws, row_idx)
            if 'シャツ' in row_text or 'パンツ' in row_text:
                header_row = row_idx
                break

        if not header_row:
            return uniforms

        # GK/FP行を探す
        for row_idx in range(header_row + 1, min(header_row + 10, ws.max_row + 1)):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, 15)]
            row_values = [str(v).strip() if v else "" for v in row_values]

            player_type = None
            if any('GK' in v or 'Ｇ' in v for v in row_values):
                player_type = "GK"
            elif any('FP' in v or 'Ｆ' in v for v in row_values):
                player_type = "FP"

            if player_type:
                # 色情報を抽出（正: 前半、副: 後半）
                colors = [v for v in row_values if v and v not in ['GK', 'FP', 'Ｇ　Ｋ', 'Ｆ　Ｐ']]

                if len(colors) >= 3:
                    # 正ユニフォーム
                    uniforms.append(ParsedUniform(
                        player_type=player_type,
                        uniform_type="primary",
                        shirt_color=colors[0] if len(colors) > 0 else None,
                        pants_color=colors[1] if len(colors) > 1 else None,
                        socks_color=colors[2] if len(colors) > 2 else None,
                    ))

                if len(colors) >= 6:
                    # 副ユニフォーム
                    uniforms.append(ParsedUniform(
                        player_type=player_type,
                        uniform_type="secondary",
                        shirt_color=colors[3] if len(colors) > 3 else None,
                        pants_color=colors[4] if len(colors) > 4 else None,
                        socks_color=colors[5] if len(colors) > 5 else None,
                    ))

        return uniforms

    def _parse_players(self, ws: Worksheet, start_row: int) -> List[ParsedPlayer]:
        """選手一覧を抽出（2列構成対応）"""
        players = []

        # ヘッダー行を探して列位置を特定
        header_row = None
        for row_idx in range(start_row, min(start_row + 5, ws.max_row + 1)):
            row_text = self._get_row_text(ws, row_idx)
            if '氏名' in row_text or 'Pos' in row_text or 'ポジション' in row_text:
                header_row = row_idx
                break

        if not header_row:
            return players

        # 列レイアウトを検出
        layout = self._detect_column_layout(ws, header_row)

        # データ行をパース
        empty_count = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, 15)]

            # 空行判定
            if not any(row):
                empty_count += 1
                if empty_count >= 3:
                    break
                continue
            empty_count = 0

            # 左列の選手
            left_player = self._extract_player_from_row(row, layout.get('left', {}), row_idx)
            if left_player:
                players.append(left_player)

            # 右列の選手
            right_player = self._extract_player_from_row(row, layout.get('right', {}), row_idx)
            if right_player:
                players.append(right_player)

        return players

    def _detect_column_layout(self, ws: Worksheet, header_row: int) -> Dict[str, Dict[str, int]]:
        """ヘッダー行から列位置を自動検出"""
        header = [ws.cell(row=header_row, column=c).value for c in range(1, 20)]
        header = [str(h).lower() if h else "" for h in header]

        layout = {'left': {}, 'right': {}}

        # 「背番号」が出現する位置で左右を分割
        split_col = None
        for idx, val in enumerate(header):
            if '背番号' in val or 'no' == val:
                split_col = idx
                break

        if split_col is None:
            # 中央で分割
            non_empty = [i for i, h in enumerate(header) if h]
            if non_empty:
                split_col = non_empty[len(non_empty) // 2]
            else:
                split_col = 5

        # 左列のマッピング
        for idx, val in enumerate(header[:split_col]):
            if 'pos' in val or 'ポジション' in val:
                layout['left']['position'] = idx
            elif '氏名' in val or '名' in val:
                layout['left']['name'] = idx
            elif '前所属' in val or '出身' in val:
                layout['left']['previous_team'] = idx
            elif '身長' in val:
                layout['left']['height'] = idx
            elif '学年' in val:
                layout['left']['grade'] = idx

        # 右列のマッピング
        for idx, val in enumerate(header[split_col:], split_col):
            if '背番号' in val or 'no' == val:
                layout['right']['number'] = idx
            elif '氏名' in val or '名' in val:
                layout['right']['name'] = idx
            elif '前所属' in val or '出身' in val:
                layout['right']['previous_team'] = idx
            elif '身長' in val:
                layout['right']['height'] = idx
            elif '学年' in val:
                layout['right']['grade'] = idx

        # デフォルト値を設定（列位置が見つからない場合）
        if 'name' not in layout['left']:
            layout['left'] = {'position': 0, 'name': 1, 'previous_team': 2, 'height': 3, 'grade': 4}
        if 'name' not in layout['right']:
            layout['right'] = {'number': 5, 'name': 6, 'previous_team': 7, 'height': 8, 'grade': 9}

        return layout

    def _extract_player_from_row(
        self, row: List, layout: Dict[str, int], row_number: int
    ) -> Optional[ParsedPlayer]:
        """行データから選手情報を抽出"""
        if not layout:
            return None

        name_idx = layout.get('name')
        if name_idx is None or name_idx >= len(row):
            return None

        name = self._clean_string(row[name_idx])
        if not name or len(name) < 2:
            return None

        player = ParsedPlayer(
            name=name,
            row_number=row_number
        )

        # 背番号
        if 'number' in layout:
            player.number = self._parse_int(row[layout['number']] if layout['number'] < len(row) else None)

        # ポジション
        if 'position' in layout:
            player.position = self._normalize_position(
                row[layout['position']] if layout['position'] < len(row) else None
            )

        # 身長
        if 'height' in layout:
            player.height = self._parse_int(row[layout['height']] if layout['height'] < len(row) else None)

        # 学年
        if 'grade' in layout:
            player.grade = self._parse_grade(row[layout['grade']] if layout['grade'] < len(row) else None)

        # 前所属
        if 'previous_team' in layout:
            player.previous_team = self._clean_string(
                row[layout['previous_team']] if layout['previous_team'] < len(row) else None
            )

        return player

    def _parse_simple_format(self, ws: Worksheet) -> List[ParsedPlayer]:
        """シンプル形式（背番号、氏名、学年等が1行ずつ）のパース"""
        players = []

        # ヘッダー行を探す
        header_row = None
        for row_idx in range(1, min(ws.max_row + 1, 20)):
            row_text = self._get_row_text(ws, row_idx)
            if '氏名' in row_text or '名前' in row_text or 'name' in row_text.lower():
                header_row = row_idx
                break

        if not header_row:
            return players

        # ヘッダーから列位置を検出
        header = [ws.cell(row=header_row, column=c).value for c in range(1, 15)]
        header = [str(h).lower() if h else "" for h in header]

        layout = {}
        for idx, val in enumerate(header):
            if '背番号' in val or 'no' == val or 'number' in val:
                layout['number'] = idx
            elif '氏名' in val or '名前' in val or 'name' in val:
                layout['name'] = idx
            elif 'フリガナ' in val or 'ふりがな' in val or 'kana' in val:
                layout['name_kana'] = idx
            elif '学年' in val or 'grade' in val:
                layout['grade'] = idx
            elif 'ポジション' in val or 'pos' in val:
                layout['position'] = idx
            elif '身長' in val or 'height' in val:
                layout['height'] = idx

        if 'name' not in layout:
            return players

        # データ行をパース
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, 15)]

            name = self._clean_string(row[layout['name']] if layout['name'] < len(row) else None)
            if not name:
                continue

            player = ParsedPlayer(name=name, row_number=row_idx)

            if 'number' in layout:
                player.number = self._parse_int(row[layout['number']])
            if 'name_kana' in layout:
                player.name_kana = self._clean_string(row[layout['name_kana']])
            if 'grade' in layout:
                player.grade = self._parse_grade(row[layout['grade']])
            if 'position' in layout:
                player.position = self._normalize_position(row[layout['position']])
            if 'height' in layout:
                player.height = self._parse_int(row[layout['height']])

            players.append(player)

        return players

    def _clean_string(self, value) -> Optional[str]:
        """文字列をクリーニング"""
        if value is None:
            return None
        s = str(value).strip()
        if s.lower() in ('nan', 'none', ''):
            return None
        # 全角スペースを半角に
        s = s.replace('　', ' ')
        # 連続スペースを1つに
        s = re.sub(r'\s+', ' ', s)
        return s.strip() if s else None

    def _parse_int(self, value) -> Optional[int]:
        """整数値に変換"""
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    def _parse_grade(self, value) -> Optional[int]:
        """学年を数値に変換"""
        if value is None:
            return None
        try:
            s = str(value)
            # "3年" や "3" から数値を抽出
            match = re.search(r'(\d)', s)
            if match:
                grade = int(match.group(1))
                if 1 <= grade <= 3:
                    return grade
        except (ValueError, TypeError):
            pass
        return None

    def _normalize_position(self, value) -> Optional[str]:
        """ポジションを正規化"""
        if value is None:
            return None
        s = str(value).upper().strip()
        return self.POSITION_MAP.get(s)


def create_players_from_parse_result(
    parse_result: ParseResult,
    team_id: int
) -> Tuple[List[Player], List[Dict]]:
    """パース結果からPlayerモデルを作成"""
    players = []
    errors = []

    for idx, parsed in enumerate(parse_result.players):
        try:
            player = Player(
                team_id=team_id,
                number=parsed.number,
                name=parsed.name,
                name_kana=parsed.name_kana,
                position=parsed.position,
                height=parsed.height,
                grade=parsed.grade,
                previous_team=parsed.previous_team,
                is_active=True,
            )
            # 正規化名を設定
            player.update_normalized_name()
            players.append(player)
        except Exception as e:
            errors.append({
                "row": parsed.row_number,
                "field": "player",
                "type": "error",
                "message": f"選手データの作成に失敗: {str(e)}"
            })

    return players, errors


def create_staff_from_parse_result(
    parse_result: ParseResult,
    team_id: int
) -> Tuple[List[Staff], List[Dict]]:
    """パース結果からStaffモデルを作成"""
    staff_list = []
    errors = []

    for idx, parsed in enumerate(parse_result.staff):
        try:
            staff = Staff(
                team_id=team_id,
                name=parsed.name,
                role=parsed.role,
                phone=parsed.phone,
                email=parsed.email,
                is_primary=(parsed.role == "監督"),
            )
            staff_list.append(staff)
        except Exception as e:
            errors.append({
                "row": idx,
                "field": "staff",
                "type": "error",
                "message": f"スタッフデータの作成に失敗: {str(e)}"
            })

    return staff_list, errors


def create_uniforms_from_parse_result(
    parse_result: ParseResult,
    team_id: int
) -> Tuple[List[TeamUniform], List[Dict]]:
    """パース結果からTeamUniformモデルを作成"""
    uniforms = []
    errors = []

    for idx, parsed in enumerate(parse_result.uniforms):
        try:
            uniform = TeamUniform(
                team_id=team_id,
                player_type=parsed.player_type,
                uniform_type=parsed.uniform_type,
                shirt_color=parsed.shirt_color,
                pants_color=parsed.pants_color,
                socks_color=parsed.socks_color,
            )
            uniforms.append(uniform)
        except Exception as e:
            errors.append({
                "row": idx,
                "field": "uniform",
                "type": "error",
                "message": f"ユニフォームデータの作成に失敗: {str(e)}"
            })

    return uniforms, errors
