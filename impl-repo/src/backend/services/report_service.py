"""
ReportService - 報告書生成サービス

PDF/Excel形式で試合報告書を生成する

このサービスは reports モジュールの DailyReportGenerator と
FinalResultPDFGenerator を使用して高品質なPDFを生成します。
"""

import io
import os
from datetime import date
from typing import Optional, List
from sqlalchemy.orm import Session, joinedload

from ..models.tournament import Tournament
from ..models.match import Match, MatchStage, MatchStatus
from ..models.venue import Venue
from ..models.goal import Goal

# 新しいレポート生成クラスをインポート
from ..reports import DailyReportGenerator, FinalResultPDFGenerator


class ReportService:
    """報告書生成サービス"""

    def __init__(self, db: Session):
        self.db = db

    def _get_matches(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> List[Match]:
        """報告書用の試合データを取得"""
        query = (
            self.db.query(Match)
            .options(
                joinedload(Match.home_team),
                joinedload(Match.away_team),
                joinedload(Match.venue),
                joinedload(Match.goals).joinedload(Goal.team),
            )
            .filter(
                Match.tournament_id == tournament_id,
                Match.match_date == target_date,
                Match.status == MatchStatus.completed,
                Match.stage != MatchStage.training,
            )
        )

        if venue_id:
            query = query.filter(Match.venue_id == venue_id)

        return query.order_by(Match.venue_id, Match.match_order).all()

    def _get_tournament(self, tournament_id: int) -> Tournament:
        """大会情報を取得"""
        tournament = self.db.query(Tournament).filter(Tournament.id == tournament_id).first()
        if not tournament:
            raise ValueError(f"大会が見つかりません (ID: {tournament_id})")
        return tournament

    def generate_pdf(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None,
        day_number: int = 1,
        recipient: Optional[str] = None
    ) -> io.BytesIO:
        """
        PDF形式の日次報告書を生成

        DailyReportGenerator を使用して高品質なPDFを生成します。

        Args:
            tournament_id: 大会ID
            target_date: 対象日付
            venue_id: 会場ID（指定時はその会場のみ、省略時は全会場）
            day_number: 大会第何日目か（デフォルト: 1）
            recipient: 送信先（省略時は大会設定から取得）

        Returns:
            PDF内容を含むBytesIO
        """
        try:
            generator = DailyReportGenerator(self.db)
            return generator.generate(
                tournament_id=tournament_id,
                target_date=target_date,
                day_number=day_number,
                recipient=recipient
            )
        except ImportError:
            # ReportLabがない場合はシンプルなPDFにフォールバック
            return self._generate_simple_pdf(tournament_id, target_date, venue_id)

    def generate_daily_report_pdf(
        self,
        tournament_id: int,
        target_date: date,
        day_number: int = 1,
        recipient: Optional[str] = None
    ) -> io.BytesIO:
        """
        日次報告書PDFを生成（generate_pdfのエイリアス）

        Args:
            tournament_id: 大会ID
            target_date: 対象日付
            day_number: 大会第何日目か
            recipient: 送信先

        Returns:
            PDF内容を含むBytesIO
        """
        return self.generate_pdf(
            tournament_id=tournament_id,
            target_date=target_date,
            day_number=day_number,
            recipient=recipient
        )

    def generate_final_result_pdf(
        self,
        tournament_id: int,
        target_date: Optional[date] = None
    ) -> io.BytesIO:
        """
        最終結果報告書PDFを生成

        FinalResultPDFGenerator を使用して最終順位、決勝トーナメント結果、
        優秀選手、研修試合結果を含むPDFを生成します。

        Args:
            tournament_id: 大会ID
            target_date: 最終日の日付（省略時は大会終了日）

        Returns:
            PDF内容を含むBytesIO
        """
        generator = FinalResultPDFGenerator(self.db)
        return generator.generate(
            tournament_id=tournament_id,
            target_date=target_date
        )

    def generate_pdf_legacy(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """
        PDF形式の報告書を生成（レガシー実装）

        ReportLabのcanvasを直接使用したシンプルなPDF生成。
        後方互換性のために残しています。新規コードではgenerate_pdfを使用してください。
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.units import mm
            from reportlab.pdfgen import canvas
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.ttfonts import TTFont
        except ImportError:
            return self._generate_simple_pdf(tournament_id, target_date, venue_id)

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        buffer = io.BytesIO()

        c = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # 日本語フォントの登録（環境変数FONT_PATHで指定可能）
        font_name = 'Helvetica'
        font_paths = [
            os.getenv('FONT_PATH'),  # 環境変数で指定されたパス
            'C:/Windows/Fonts/msgothic.ttc',  # Windows
            '/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc',  # macOS
            '/usr/share/fonts/truetype/fonts-japanese-gothic.ttf',  # Linux (Debian/Ubuntu)
            '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc',  # Linux (Noto)
        ]
        for font_path in font_paths:
            if font_path and os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('Gothic', font_path))
                    font_name = 'Gothic'
                    break
                except Exception:
                    continue

        # ヘッダー
        c.setFont(font_name, 16)
        c.drawString(30 * mm, height - 20 * mm, f"{tournament.name}")
        c.setFont(font_name, 12)
        c.drawString(30 * mm, height - 30 * mm, f"試合結果報告書 - {target_date.isoformat()}")

        if venue_id:
            venue = self.db.query(Venue).filter(Venue.id == venue_id).first()
            if venue:
                c.drawString(30 * mm, height - 40 * mm, f"会場: {venue.name}")

        # 試合情報
        y_position = height - 60 * mm
        c.setFont(font_name, 10)

        for match in matches:
            if y_position < 40 * mm:
                c.showPage()
                y_position = height - 20 * mm
                c.setFont(font_name, 10)

            home = match.home_team.short_name or match.home_team.name if match.home_team else "TBD"
            away = match.away_team.short_name or match.away_team.name if match.away_team else "TBD"
            time_str = match.match_time.strftime("%H:%M") if match.match_time else ""

            h1 = match.home_score_half1 or 0
            h2 = match.home_score_half2 or 0
            a1 = match.away_score_half1 or 0
            a2 = match.away_score_half2 or 0

            c.drawString(30 * mm, y_position, f"第{match.match_order}試合 {time_str}")
            y_position -= 6 * mm

            score_line = f"{home} {h1+h2} - {a1+a2} {away}"
            c.drawString(35 * mm, y_position, score_line)
            y_position -= 5 * mm

            c.drawString(35 * mm, y_position, f"(前半 {h1}-{a1}、後半 {h2}-{a2})")

            if match.has_penalty_shootout:
                y_position -= 5 * mm
                c.drawString(35 * mm, y_position, f"PK {match.home_pk}-{match.away_pk}")

            y_position -= 5 * mm

            # 得点経過
            for goal in sorted(match.goals, key=lambda g: (g.half, g.minute)):
                half_text = "前半" if goal.half == 1 else "後半"
                goal_text = f"{half_text}{goal.minute}分 {goal.team.name} {goal.player_name}"
                if goal.is_own_goal:
                    goal_text += "(OG)"
                if goal.is_penalty:
                    goal_text += "(PK)"

                c.drawString(40 * mm, y_position, goal_text)
                y_position -= 4 * mm

            y_position -= 8 * mm

        # フッター
        c.setFont(font_name, 8)
        c.drawString(30 * mm, 15 * mm, "浦和カップ運営事務局")

        c.save()
        buffer.seek(0)

        return buffer

    def _generate_simple_pdf(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """シンプルなテキストベースのPDFを生成（ReportLabがない場合）"""
        buffer = io.BytesIO()

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        # 簡易PDFヘッダー
        pdf_content = b"%PDF-1.4\n"
        pdf_content += b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
        pdf_content += b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
        pdf_content += b"3 0 obj<</Type/Page/MediaBox[0 0 595 842]/Parent 2 0 R/Resources<<>>/Contents 4 0 R>>endobj\n"

        text = f"Match Report - {target_date.isoformat()}\n\n"
        for match in matches:
            home = match.home_team.name if match.home_team else "TBD"
            away = match.away_team.name if match.away_team else "TBD"
            score = f"{match.home_score_total or 0}-{match.away_score_total or 0}"
            text += f"{match.match_order}. {home} {score} {away}\n"

        stream = f"4 0 obj<</Length {len(text)}>>stream\n{text}\nendstream endobj\n"
        pdf_content += stream.encode('latin-1', errors='replace')
        pdf_content += b"xref\n0 5\n0000000000 65535 f \n"
        pdf_content += b"trailer<</Size 5/Root 1 0 R>>\nstartxref\n0\n%%EOF"

        buffer.write(pdf_content)
        buffer.seek(0)

        return buffer

    def generate_excel(
        self,
        tournament_id: int,
        target_date: date,
        venue_id: Optional[int] = None
    ) -> io.BytesIO:
        """Excel形式の報告書を生成"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            raise ImportError("openpyxlがインストールされていません。pip install openpyxl を実行してください。")

        tournament = self._get_tournament(tournament_id)
        matches = self._get_matches(tournament_id, target_date, venue_id)

        wb = Workbook()
        ws = wb.active
        ws.title = "試合結果"

        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # ヘッダー
        ws.merge_cells('A1:G1')
        ws['A1'] = tournament.name
        ws['A1'].font = header_font

        ws.merge_cells('A2:G2')
        ws['A2'] = f"試合結果報告書 - {target_date.isoformat()}"
        ws['A2'].font = subheader_font

        if venue_id:
            venue = self.db.query(Venue).filter(Venue.id == venue_id).first()
            if venue:
                ws.merge_cells('A3:G3')
                ws['A3'] = f"会場: {venue.name}"

        # テーブルヘッダー
        row = 5
        headers = ['試合', '時刻', 'ホーム', '前半', '後半', '合計', 'アウェイ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = subheader_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 8
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 20

        # 試合データ
        for match in matches:
            row += 1

            home = match.home_team.short_name or match.home_team.name if match.home_team else "TBD"
            away = match.away_team.short_name or match.away_team.name if match.away_team else "TBD"
            time_str = match.match_time.strftime("%H:%M") if match.match_time else ""

            h1 = match.home_score_half1 or 0
            h2 = match.home_score_half2 or 0
            a1 = match.away_score_half1 or 0
            a2 = match.away_score_half2 or 0

            values = [
                match.match_order,
                time_str,
                home,
                f"{h1}-{a1}",
                f"{h2}-{a2}",
                f"{h1+h2}-{a1+a2}",
                away,
            ]

            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')

            if match.has_penalty_shootout:
                row += 1
                ws.cell(row=row, column=6, value=f"PK {match.home_pk}-{match.away_pk}")

            for goal in sorted(match.goals, key=lambda g: (g.half, g.minute)):
                row += 1
                half_text = "前半" if goal.half == 1 else "後半"
                goal_text = f"{half_text}{goal.minute}分 {goal.team.name} {goal.player_name}"
                if goal.is_own_goal:
                    goal_text += "(OG)"
                if goal.is_penalty:
                    goal_text += "(PK)"

                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=7)
                ws.cell(row=row, column=3, value=goal_text).font = normal_font

            row += 1

        row += 2
        ws.cell(row=row, column=1, value="浦和カップ運営事務局")

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer
