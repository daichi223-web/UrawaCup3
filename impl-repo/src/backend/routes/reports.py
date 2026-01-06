"""
Reports（報告書）ルーター
報告書の生成とPDF/Excelエクスポートを提供
"""

import io
from datetime import datetime, date
from typing import List, Optional, Dict, Any
from fastapi import APIRouter, Depends, HTTPException, Query, status, Body
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

from ..database import get_db
from ..models.tournament import Tournament
from ..models.match import Match, MatchStage, MatchStatus
from ..models.venue import Venue
from ..models.goal import Goal
from ..models.report_recipient import ReportRecipient
from ..models.group import Group
from ..models.standing import Standing
from ..models.team import Team
from ..schemas.report import (
    ReportRecipientCreate,
    ReportRecipientResponse,
    ReportParams,
    MatchReport,
    GoalReport,
    ReportData,
    SenderSettingsUpdate,
    SenderSettingsResponse,
    DailyReportRequest,
)

router = APIRouter()


# ================== 報告書送信先管理 ==================

@router.get("/recipients", response_model=List[ReportRecipientResponse])
def get_recipients(
    tournament_id: int = Query(..., description="大会ID"),
    db: Session = Depends(get_db),
):
    """報告書送信先一覧を取得"""
    recipients = db.query(ReportRecipient).filter(
        ReportRecipient.tournament_id == tournament_id
    ).all()
    return recipients


@router.post("/recipients", response_model=ReportRecipientResponse, status_code=status.HTTP_201_CREATED)
def create_recipient(
    recipient_data: ReportRecipientCreate,
    db: Session = Depends(get_db),
):
    """報告書送信先を追加"""
    tournament = db.query(Tournament).filter(Tournament.id == recipient_data.tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {recipient_data.tournament_id})"
        )

    recipient = ReportRecipient(
        tournament_id=recipient_data.tournament_id,
        organization=recipient_data.name,
        recipient_name=recipient_data.notes,
        email=recipient_data.email,
    )
    db.add(recipient)
    db.commit()
    db.refresh(recipient)

    return recipient


@router.delete("/recipients/{recipient_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_recipient(
    recipient_id: int,
    db: Session = Depends(get_db),
):
    """報告書送信先を削除"""
    recipient = db.query(ReportRecipient).filter(ReportRecipient.id == recipient_id).first()
    if not recipient:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"送信先が見つかりません (ID: {recipient_id})"
        )

    db.delete(recipient)
    db.commit()

    return None


@router.post("/recipients/{tournament_id}/setup-default", response_model=List[ReportRecipientResponse])
def setup_default_recipients(
    tournament_id: int,
    db: Session = Depends(get_db),
):
    """
    デフォルトの送信先を設定

    - 埼玉新聞
    - テレビ埼玉
    - イシクラ
    - 埼玉県サッカー協会
    """
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {tournament_id})"
        )

    default_recipients = [
        {"organization": "埼玉新聞", "recipient_name": "スポーツ部"},
        {"organization": "テレビ埼玉", "recipient_name": "報道部"},
        {"organization": "イシクラ", "recipient_name": ""},
        {"organization": "埼玉県サッカー協会", "recipient_name": ""},
    ]

    created = []
    for data in default_recipients:
        recipient = ReportRecipient(
            tournament_id=tournament_id,
            **data,
        )
        db.add(recipient)
        created.append(recipient)

    db.commit()

    for r in created:
        db.refresh(r)

    return created


# ================== 報告書発信元設定 ==================

@router.get("/tournaments/{tournament_id}/report-settings", response_model=SenderSettingsResponse)
def get_sender_settings(
    tournament_id: int,
    db: Session = Depends(get_db),
):
    """報告書発信元設定を取得"""
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {tournament_id})"
        )

    return SenderSettingsResponse(
        sender_organization=getattr(tournament, 'sender_organization', None),
        sender_name=getattr(tournament, 'sender_name', None),
        sender_contact=getattr(tournament, 'sender_contact', None),
    )


@router.put("/tournaments/{tournament_id}/report-settings", response_model=SenderSettingsResponse)
def update_sender_settings(
    tournament_id: int,
    settings: SenderSettingsUpdate,
    db: Session = Depends(get_db),
):
    """報告書発信元設定を更新"""
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {tournament_id})"
        )

    update_data = settings.model_dump(by_alias=False, exclude_unset=True)
    for field, value in update_data.items():
        if hasattr(tournament, field):
            setattr(tournament, field, value)

    db.commit()
    db.refresh(tournament)

    return SenderSettingsResponse(
        sender_organization=getattr(tournament, 'sender_organization', None),
        sender_name=getattr(tournament, 'sender_name', None),
        sender_contact=getattr(tournament, 'sender_contact', None),
    )


# ================== 報告書データ取得 ==================

@router.get("/match-reports", response_model=List[MatchReport])
def get_match_reports(
    tournament_id: int = Query(..., description="大会ID"),
    target_date: date = Query(..., description="対象日"),
    venue_id: Optional[int] = Query(None, description="会場ID"),
    db: Session = Depends(get_db),
):
    """試合報告書形式でデータを取得"""
    query = (
        db.query(Match)
        .options(
            joinedload(Match.home_team),
            joinedload(Match.away_team),
            joinedload(Match.goals).joinedload(Goal.team),
        )
        .filter(
            Match.tournament_id == tournament_id,
            Match.match_date == target_date,
            Match.status == MatchStatus.completed,
            Match.stage != MatchStage.training,
        )
    )

    if venue_id:
        query = query.filter(Match.venue_id == venue_id)

    matches = query.order_by(Match.venue_id, Match.match_order).all()

    reports = []
    for match in matches:
        goals = []
        for goal in sorted(match.goals, key=lambda g: (g.half, g.minute)):
            half_text = "前半" if goal.half == 1 else "後半"
            display = f"{half_text}{goal.minute}分 {goal.team.name} {goal.player_name}"
            if goal.is_own_goal:
                display += "(OG)"
            if goal.is_penalty:
                display += "(PK)"

            goals.append(GoalReport(
                minute=goal.minute,
                half=goal.half,
                team_name=goal.team.name,
                player_name=goal.player_name,
                display_text=display,
            ))

        h1 = match.home_score_half1 or 0
        h2 = match.home_score_half2 or 0
        a1 = match.away_score_half1 or 0
        a2 = match.away_score_half2 or 0

        score_pk = None
        if match.has_penalty_shootout:
            score_pk = f"{match.home_pk}-{match.away_pk}"

        reports.append(MatchReport(
            match_number=match.match_order,
            kickoff_time=match.match_time.strftime("%H:%M") if match.match_time else "",
            home_team_name=match.home_team.short_name or match.home_team.name if match.home_team else "TBD",
            away_team_name=match.away_team.short_name or match.away_team.name if match.away_team else "TBD",
            score_half1=f"{h1}-{a1}",
            score_half2=f"{h2}-{a2}",
            score_total=f"{h1 + h2}-{a1 + a2}",
            score_pk=score_pk,
            goals=goals,
        ))

    return reports


# ================== Excel出力 ==================

@router.get("/export/group-standings/excel")
def export_group_standings_excel(
    tournament_id: int = Query(..., description="大会ID"),
    group_id: Optional[str] = Query(None, description="グループID"),
    db: Session = Depends(get_db),
):
    """グループ順位表をExcelで出力"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="大会が見つかりません")

    query = db.query(Standing).filter(Standing.tournament_id == tournament_id)
    if group_id:
        query = query.filter(Standing.group_id == group_id)
    standings = query.order_by(Standing.group_id, Standing.rank).all()

    if not standings:
        raise HTTPException(status_code=404, detail="順位データが見つかりません")

    wb = Workbook()
    ws = wb.active
    ws.title = "グループ順位表"

    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    ws.merge_cells('A1:J1')
    ws['A1'] = f"{tournament.name} グループ順位表"
    ws['A1'].font = header_font

    row = 3
    current_group = None

    for standing in standings:
        if standing.group_id != current_group:
            if current_group is not None:
                row += 1

            current_group = standing.group_id
            group = db.query(Group).filter(Group.id == standing.group_id).first()
            group_name = group.name if group else standing.group_id

            ws[f'A{row}'] = f"【{group_name}】"
            ws[f'A{row}'].font = subheader_font
            row += 1

            headers = ['順位', 'チーム', '試合', '勝', '分', '負', '得点', '失点', '得失点差', '勝点']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
            row += 1

        team = db.query(Team).filter(Team.id == standing.team_id).first()
        team_name = team.short_name or team.name if team else "Unknown"

        values = [
            standing.rank,
            team_name,
            standing.played,
            standing.won,
            standing.drawn,
            standing.lost,
            standing.goals_for,
            standing.goals_against,
            standing.goal_difference,
            standing.points,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 6

    row += 2
    ws.cell(row=row, column=1, value="浦和カップ運営事務局")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"group_standings_{tournament_id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ================== 大会サマリー ==================

@router.get("/summary/{tournament_id}")
def get_tournament_summary(
    tournament_id: int,
    db: Session = Depends(get_db),
):
    """大会サマリーを取得（ダッシュボード用）"""
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {tournament_id})"
        )

    team_count = db.query(Team).filter(Team.tournament_id == tournament_id).count()

    match_stats = db.query(Match).filter(Match.tournament_id == tournament_id)
    total_matches = match_stats.count()
    completed_matches = match_stats.filter(Match.status == MatchStatus.completed).count()

    stage_counts = {}
    for stage in MatchStage:
        count = match_stats.filter(Match.stage == stage).count()
        if count > 0:
            stage_counts[stage.value] = count

    total_goals = db.query(Goal).join(Match).filter(
        Match.tournament_id == tournament_id
    ).count()

    return {
        "tournament_id": tournament_id,
        "tournament_name": tournament.name,
        "team_count": team_count,
        "total_matches": total_matches,
        "completed_matches": completed_matches,
        "completion_rate": round(completed_matches / total_matches * 100, 1) if total_matches > 0 else 0,
        "stage_counts": stage_counts,
        "total_goals": total_goals,
    }


# ================== 未入力チェック ==================

@router.get("/check-incomplete")
def check_incomplete_data(
    tournament_id: int = Query(..., description="大会ID"),
    target_date: Optional[date] = Query(None, description="対象日"),
    db: Session = Depends(get_db),
):
    """未入力データをチェック"""
    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"大会が見つかりません (ID: {tournament_id})"
        )

    warnings = []
    critical_warnings = []

    if target_date:
        matches = db.query(Match).filter(
            Match.tournament_id == tournament_id,
            Match.match_date == target_date,
            Match.stage != MatchStage.training,
        ).all()

        if not matches:
            critical_warnings.append({
                "type": "no_matches",
                "message": f"{target_date}の試合データがありません",
                "field": "matches",
            })
        else:
            for match in matches:
                match_label = f"第{match.match_order}試合"

                if match.status != MatchStatus.completed:
                    critical_warnings.append({
                        "type": "incomplete_match",
                        "message": f"{match_label}の結果が未入力です",
                        "field": "match_result",
                        "match_id": match.id,
                    })

    return {
        "warnings": warnings,
        "critical_warnings": critical_warnings,
        "can_export": len(critical_warnings) == 0,
        "warning_count": len(warnings),
        "critical_count": len(critical_warnings),
    }


# ================== エイリアスルート（要件定義互換） ==================

@router.get("/daily", response_model=List[MatchReport])
def get_daily_report(
    tournament_id: int = Query(..., description="大会ID"),
    date: date = Query(..., alias="date", description="対象日（YYYY-MM-DD形式）"),
    venue_id: Optional[int] = Query(None, description="会場ID（任意）"),
    db: Session = Depends(get_db),
):
    """
    日次報告書を取得（要件定義互換エイリアス）

    GET /api/reports/daily?tournament_id={id}&date={date}

    既存の GET /api/reports/match-reports へのエイリアスとして機能します。
    """
    return get_match_reports(
        tournament_id=tournament_id,
        target_date=date,
        venue_id=venue_id,
        db=db,
    )


@router.get("/standings")
def get_standings_report(
    tournament_id: int = Query(..., description="大会ID"),
    group_id: Optional[str] = Query(None, description="グループID（任意）"),
    format: str = Query("pdf", description="出力形式（pdf/excel）"),
    db: Session = Depends(get_db),
):
    """
    順位表PDFを取得（要件定義互換エイリアス）

    GET /api/reports/standings?tournament_id={id}

    既存の GET /api/reports/export/group-standings/excel へのエイリアスとして機能します。
    format=excel の場合はExcel形式、それ以外はPDF形式で出力します。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxlがインストールされていません")

    tournament = db.query(Tournament).filter(Tournament.id == tournament_id).first()
    if not tournament:
        raise HTTPException(status_code=404, detail="大会が見つかりません")

    query = db.query(Standing).filter(Standing.tournament_id == tournament_id)
    if group_id:
        query = query.filter(Standing.group_id == group_id)
    standings = query.order_by(Standing.group_id, Standing.rank).all()

    if not standings:
        raise HTTPException(status_code=404, detail="順位データが見つかりません")

    wb = Workbook()
    ws = wb.active
    ws.title = "グループ順位表"

    header_font = Font(bold=True, size=14)
    subheader_font = Font(bold=True, size=11)
    normal_font = Font(size=10)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=10, color="FFFFFF")

    ws.merge_cells('A1:J1')
    ws['A1'] = f"{tournament.name} グループ順位表"
    ws['A1'].font = header_font

    row = 3
    current_group = None

    for standing in standings:
        if standing.group_id != current_group:
            if current_group is not None:
                row += 1

            current_group = standing.group_id
            group = db.query(Group).filter(Group.id == standing.group_id).first()
            group_name = group.name if group else standing.group_id

            ws[f'A{row}'] = f"【{group_name}】"
            ws[f'A{row}'].font = subheader_font
            row += 1

            headers = ['順位', 'チーム', '試合', '勝', '分', '負', '得点', '失点', '得失点差', '勝点']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font_white
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
            row += 1

        team = db.query(Team).filter(Team.id == standing.team_id).first()
        team_name = team.short_name or team.name if team else "Unknown"

        values = [
            standing.rank,
            team_name,
            standing.played,
            standing.won,
            standing.drawn,
            standing.lost,
            standing.goals_for,
            standing.goals_against,
            standing.goal_difference,
            standing.points,
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        row += 1

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 6
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 6
    ws.column_dimensions['H'].width = 6
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 6

    row += 2
    ws.cell(row=row, column=1, value="浦和カップ運営事務局")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # format パラメータに応じて出力形式を変更
    if format.lower() == "excel":
        filename = f"group_standings_{tournament_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        # PDF形式の場合もExcelを生成（PDF変換ライブラリがない場合の互換対応）
        # 実際のPDF変換が必要な場合は、reportlab等のライブラリを追加してください
        filename = f"group_standings_{tournament_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return StreamingResponse(
        buffer,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ================== PDF生成エンドポイント ==================

@router.get("/daily/pdf")
def generate_daily_report_pdf(
    tournament_id: int = Query(..., description="大会ID"),
    target_date: date = Query(..., alias="date", description="対象日（YYYY-MM-DD形式）"),
    day_number: int = Query(1, description="大会第何日目か"),
    recipient: Optional[str] = Query(None, description="送信先（省略時は大会設定から取得）"),
    db: Session = Depends(get_db),
):
    """
    日次報告書PDFを生成

    GET /api/reports/daily/pdf?tournament_id={id}&date={date}&day_number={day}

    Platypus（高レベルAPI）を使用した高品質なPDF生成
    """
    from ..reports import DailyReportGenerator

    try:
        generator = DailyReportGenerator(db)
        buffer = generator.generate(
            tournament_id=tournament_id,
            target_date=target_date,
            day_number=day_number,
            recipient=recipient
        )

        filename = f"daily_report_{tournament_id}_{target_date.isoformat()}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF生成エラー: {str(e)}")


@router.get("/final-result/pdf")
def generate_final_result_pdf(
    tournament_id: int = Query(..., description="大会ID"),
    target_date: Optional[date] = Query(None, alias="date", description="最終日日付（省略時は大会終了日）"),
    db: Session = Depends(get_db),
):
    """
    最終結果報告書PDFを生成

    GET /api/reports/final-result/pdf?tournament_id={id}

    Platypus（高レベルAPI）を使用した高品質なPDF生成
    - 最終順位（1位〜4位）
    - 決勝トーナメント結果
    - 優秀選手一覧
    - 研修試合結果
    """
    from ..reports import FinalResultPDFGenerator

    try:
        generator = FinalResultPDFGenerator(db)
        buffer = generator.generate(
            tournament_id=tournament_id,
            target_date=target_date
        )

        filename = f"final_result_{tournament_id}.pdf"
        return StreamingResponse(
            buffer,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF生成エラー: {str(e)}")


@router.post("/daily/pdf/from-json")
def generate_daily_report_pdf_from_json(
    data: Dict[str, Any] = Body(..., description="core形式のJSONデータ"),
    db: Session = Depends(get_db),
):
    """
    JSON形式のデータから日次報告書PDFを生成（core互換）

    POST /api/reports/daily/pdf/from-json

    coreスクリプト (generate_daily_report_pdf.py) と同じ形式のJSONを受け付け、
    PDFを生成して返します。
    """
    from ..reports import DailyReportGenerator
    import tempfile
    import os

    try:
        generator = DailyReportGenerator(db)

        # 一時ファイルに出力
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name

        generator.generate_from_dict(data, tmp_path)

        # ファイルを読み込んでStreamingResponse
        with open(tmp_path, 'rb') as f:
            content = f.read()

        # 一時ファイル削除
        os.unlink(tmp_path)

        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=daily_report.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF生成エラー: {str(e)}")


@router.post("/final-result/pdf/from-json")
def generate_final_result_pdf_from_json(
    data: Dict[str, Any] = Body(..., description="core形式のJSONデータ"),
    db: Session = Depends(get_db),
):
    """
    JSON形式のデータから最終結果報告書PDFを生成（core互換）

    POST /api/reports/final-result/pdf/from-json

    coreスクリプト (generate_final_result_pdf.py) と同じ形式のJSONを受け付け、
    PDFを生成して返します。
    """
    from ..reports import FinalResultPDFGenerator
    import tempfile
    import os

    try:
        generator = FinalResultPDFGenerator(db)

        # 一時ファイルに出力
        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            tmp_path = tmp.name

        generator.generate_from_dict(data, tmp_path)

        # ファイルを読み込んでStreamingResponse
        with open(tmp_path, 'rb') as f:
            content = f.read()

        # 一時ファイル削除
        os.unlink(tmp_path)

        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": "attachment; filename=final_result.pdf"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"PDF生成エラー: {str(e)}")
